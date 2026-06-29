import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from PySide6.QtWidgets import QFileDialog, QMessageBox, QProgressDialog, QApplication
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus.flowables import HRFlowable
from reportlab.platypus import PageBreak, Frame, PageTemplate
from reportlab.platypus.flowables import Image # For PDF logo
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from PySide6.QtCore import Qt


def _make_progress(parent, title):
    if QApplication.instance() is None:
        return None

    progress = QProgressDialog(title, None, 0, 100, parent)
    progress.setWindowTitle(title)
    progress.setWindowModality(Qt.WindowModal if parent else Qt.NonModal)
    progress.setAutoClose(False)
    progress.setAutoReset(False)
    progress.setMinimumDuration(0)
    progress.setValue(0)
    QApplication.processEvents()
    return progress


def _set_progress(progress, value, label):
    if progress is None:
        return
    progress.setLabelText(f"{label} {value}%")
    progress.setValue(value)
    QApplication.processEvents()




def _report_palette():
    return {
        "accent": "2563EB",
        "accent_dark": "1B3A5C",
        "light": "E8EDF2",
        "text": "111827",
    }


def _hex_color(value):
    return colors.HexColor(f"#{value}")


def _pick(mapping, *keys, default="-"):
    for key in keys:
        if isinstance(mapping, dict) and key in mapping:
            return mapping[key]
    return default


def to_excel(parent, data):
    path, _ = QFileDialog.getSaveFileName(parent, "Export Broadsheet", f"Broadsheet_{data['meta']['class']}.xlsx", "Excel Files (*.xlsx)")
    if not path: return

    progress = None

    try:
        progress = _make_progress(parent, "Exporting Excel broadsheet...")
        wb = openpyxl.Workbook()
        _set_progress(progress, 5, "Preparing workbook")
        ws = wb.active
        ws.title = "Broadsheet"

        # Report-card style header
        meta = data['meta']
        school_profile = meta['school_profile']
        palette = _report_palette()
        last_col = len(data['subjects']) + 8
        school_name = school_profile.get('school_name', 'SCHOOL MANAGEMENT SYSTEM').upper()
        motto = school_profile.get('school_motto') or school_profile.get('motto') or ''
        address = school_profile.get('school_address') or school_profile.get('school_address', '-') or '-'
        phone = school_profile.get('school_phone') or school_profile.get('phone') or '-'
        email = school_profile.get('school_email') or school_profile.get('email') or '-'

        header_rows = [
            (school_name, 18, palette['accent_dark'], "FFFFFF"),
            (motto, 11, palette['accent_dark'], "FFFFFF"),
            (f"{address} | Tel: {phone} | Email: {email}", 10, palette['light'], palette['text']),
            ("CLASS BROADSHEET / REPORT SUMMARY", 14, palette['accent'], "FFFFFF"),
            (f"Class: {meta['class']}   Level: {meta['level']}   Exam: {meta['exam']}   Term: {meta['term']}   Year: {meta['year']}", 11, palette['light'], palette['text']),
        ]

        for text, size, fill, font_color in header_rows:
            ws.append([text])
            row_no = ws.max_row
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=last_col)
            cell = ws.cell(row=row_no, column=1)
            cell.font = Font(size=size, bold=True, color=font_color)
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_no].height = 24 if size >= 14 else 19

        ws.append([])

        # Main Broadsheet Table Header
        subjects = data['subjects']
        headers = ["Position", "Admission No", "Student Name", "Gender"] + subjects + ["Total", "Average", "Points", "Division"]
        
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=palette["accent_dark"], end_color=palette["accent_dark"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data
        total_data_rows = max(len(data['rows']), 1)
        for row_index, r in enumerate(data['rows'], start=1):
            row_vals = [r['Position'], r['Admission No'], r['Student Name'], r['Gender']]
            for s in subjects:
                row_vals.append(r['marks'][s])
            row_vals += [r['Total'], r['Average'], r['Points'], r['Division']]
            ws.append(row_vals)
            if row_index == 1 or row_index == total_data_rows or row_index % 25 == 0:
                _set_progress(progress, min(55, 10 + int((row_index / total_data_rows) * 45)), "Writing student rows")

        _set_progress(progress, 60, "Styling workbook")

        # Add a border to the main table
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        for r_idx in range(header_row, ws.max_row + 1):
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).border = thin_border

        ws.append([]) # spacer

        # 1. Class Performance Analysis
        ws.append(["CLASS PERFORMANCE ANALYSIS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        class_perf = data['class_performance']
        ws.append(["Total Students", class_perf['total_students'], "Class Average", f"{class_perf['class_average']}%"])
        ws.append(["Highest Average", f"{class_perf['highest_average']}%", "Lowest Average", f"{class_perf['lowest_average']}%"])
        ws.append(["Pass Rate", f"{class_perf['pass_rate']}%", "Fail Rate", f"{class_perf['fail_rate']}%"])
        ws.append([])

        # 2. Gender Summary
        if data['settings']['show_gender_summary']:
            ws.append(["GENDER SUMMARY"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            gender_sum = data['gender_summary']
            ws.append(["Gender", "Count"])
            ws.append(["Male", gender_sum['Male']])
            ws.append(["Female", gender_sum['Female']])
            ws.append(["Total", gender_sum['Total']])
            ws.append([])

        # 3. Division Summary
        ws.append(["DIVISION SUMMARY"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Division", "Students"])
        for div, count in data['division_summary'].items():
            ws.append([div, count])
        ws.append([])

        # 4. Top 10 Students
        ws.append(["TOP 10 STUDENTS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Position", "Admission No", "Student Name", "Average", "Division"])
        for s in data['top_students']:
            ws.append([s['position'], s['admission'], s['name'], s['average'], s['division']])
        ws.append([])

        # 5. Bottom 10 Students
        ws.append(["BOTTOM 10 STUDENTS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Position", "Admission No", "Student Name", "Average", "Division"])
        for s in data['bottom_students']:
            ws.append([s['position'], s['admission'], s['name'], s['average'], s['division']])
        ws.append([])

        # 6. Subject Performance Analysis
        ws.append(["SUBJECT PERFORMANCE ANALYSIS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Subject", "Average", "Passes", "Fails"])
        best_sub = None
        worst_sub = None
        max_avg = -1
        min_avg = 101
        for sub_name, stats in data['subject_performance'].items():
            ws.append([sub_name, stats['average'], stats['passes'], stats['fails']])
            if stats['average'] > max_avg:
                max_avg = stats['average']
                best_sub = sub_name
            if stats['average'] < min_avg:
                min_avg = stats['average']
                worst_sub = sub_name
        ws.append([f"Best Subject: {best_sub} (Avg: {max_avg})"])
        ws.append([f"Worst Subject: {worst_sub} (Avg: {min_avg})"])
        ws.append([])

        # Signatures section
        ws.append([])
        ws.append(["Academic Master Signature: ...................................................."])
        ws.append(["Head Teacher Signature: ...................................................."])
        ws.append(["School Stamp:"])

        # Style data cells
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Freeze Panes
        ws.freeze_panes = f"E{header_row + 1}"

        _set_progress(progress, 90, "Saving Excel file")
        wb.save(path)
        _set_progress(progress, 100, "Export complete")
        QMessageBox.information(parent, "Success", f"Broadsheet exported to {path}")
    except Exception as e:
        print(f"[ERROR] Broadsheet export failed: {e}")
        QMessageBox.critical(parent, "Export Error", "An unexpected error occurred during export.")
    finally:
        if progress is not None:
            progress.close()


def to_pdf(parent, data):
    path, _ = QFileDialog.getSaveFileName(parent, "Export Broadsheet PDF", f"Broadsheet_{data['meta']['class']}.pdf", "PDF Files (*.pdf)")
    if not path: return

    progress = None

    try:
        progress = _make_progress(parent, "Exporting PDF broadsheet...")
        doc = SimpleDocTemplate(path, pagesize=landscape(A4), 
                                rightMargin=0.5*inch, leftMargin=0.5*inch, 
                                topMargin=0.75*inch, bottomMargin=0.5*inch)
        
        _set_progress(progress, 8, "Preparing PDF document")
        styles = getSampleStyleSheet()
        
        # Custom styles for header/footer
        styles.add(ParagraphStyle(name='HeaderStyle', alignment=1, fontSize=10, fontName='Helvetica-Bold', spaceAfter=0))
        styles.add(ParagraphStyle(name='FooterStyle', alignment=1, fontSize=8, fontName='Helvetica', spaceAfter=0))
        styles.add(ParagraphStyle(name='TitleStyle', alignment=1, fontSize=16, fontName='Helvetica-Bold', spaceAfter=0))
        styles.add(ParagraphStyle(name='SubtitleStyle', alignment=1, fontSize=10, fontName='Helvetica', spaceAfter=0))
        styles.add(ParagraphStyle(name='SectionHeader', alignment=0, fontSize=12, fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=6))
        styles.add(ParagraphStyle(name='ReportTitle', alignment=TA_CENTER, fontSize=17, fontName='Helvetica-Bold', leading=20, textColor=colors.white))
        styles.add(ParagraphStyle(name='ReportMotto', alignment=TA_CENTER, fontSize=8, fontName='Helvetica-Oblique', leading=10, textColor=colors.white))
        styles.add(ParagraphStyle(name='ReportSmall', alignment=TA_CENTER, fontSize=8, fontName='Helvetica', leading=10))
        styles.add(ParagraphStyle(name='ReportSmallLeft', alignment=TA_LEFT, fontSize=8, fontName='Helvetica', leading=10, textColor=colors.white))
        styles.add(ParagraphStyle(name='ReportSmallRight', alignment=TA_RIGHT, fontSize=8, fontName='Helvetica', leading=10, textColor=colors.white))

        # School Profile and Settings
        meta = data['meta']
        school_profile = meta['school_profile']
        settings = data['settings']
        generated_date = meta['generated_date']
        palette = _report_palette()
        accent = _hex_color(palette['accent'])
        accent_dark = _hex_color(palette['accent_dark'])
        light = _hex_color(palette['light'])

        # Define Header and Footer for all pages
        def _header_footer(canvas, doc):
            canvas.saveState()
            header_frame = Frame(
                doc.leftMargin,
                doc.height + doc.topMargin - 1.05 * inch,
                doc.width,
                1.05 * inch,
                leftPadding=0,
                bottomPadding=0,
                rightPadding=0,
                topPadding=0,
                showBoundary=0,
            )

            school_name = school_profile.get(
                'school_name', 'SCHOOL MANAGEMENT SYSTEM'
            ).upper()
            motto = school_profile.get('school_motto') or school_profile.get('motto') or ''
            address = school_profile.get('school_address') or '-'
            phone = school_profile.get('school_phone') or '-'
            email = school_profile.get('school_email') or '-'

            center_parts = []
            if settings['show_logo'] and school_profile.get('school_logo') and os.path.exists(school_profile['school_logo']):
                try:
                    logo = Image(school_profile['school_logo'], width=0.55 * inch, height=0.55 * inch)
                    center_parts.append([logo])
                except Exception as e:
                    print(f"[WARNING] Could not load PDF logo '{school_profile['school_logo']}': {e}")

            center_parts.append([Paragraph(school_name, styles['ReportTitle'])])
            if motto:
                center_parts.append([Paragraph(motto, styles['ReportMotto'])])

            center_table = Table(center_parts, colWidths=[doc.width * 0.42])
            center_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ]))

            left = Table([
                [Paragraph('<b>CONTACT</b>', styles['ReportSmallLeft'])],
                [Paragraph(address, styles['ReportSmallLeft'])],
                [Paragraph(f'Tel: {phone}', styles['ReportSmallLeft'])],
                [Paragraph(f'Email: {email}', styles['ReportSmallLeft'])],
            ], colWidths=[doc.width * 0.29])

            right = Table([
                [Paragraph('<b>ACADEMIC REPORT</b>', styles['ReportSmallRight'])],
                [Paragraph(f"Class: {meta['class']} ({meta['level']})", styles['ReportSmallRight'])],
                [Paragraph(f"Exam: {meta['exam']}", styles['ReportSmallRight'])],
                [Paragraph(f"{meta['term']} - {meta['year']}", styles['ReportSmallRight'])],
            ], colWidths=[doc.width * 0.29])

            for panel in (left, right):
                panel.setStyle(TableStyle([
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ]))

            header = Table(
                [[left, center_table, right]],
                colWidths=[doc.width * 0.29, doc.width * 0.42, doc.width * 0.29],
            )
            header.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), accent_dark),
                ('BOX', (0, 0), (-1, -1), 1.2, accent),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))

            header_frame.addFromList([header], canvas)

            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.HexColor('#64748B'))
            canvas.drawString(doc.leftMargin, 0.3 * inch, f"Generated: {generated_date}")
            canvas.drawString(doc.width + doc.leftMargin - 0.5 * inch, 0.3 * inch, f"Page {doc.page}")
            canvas.restoreState()

        # Watermark
        def draw_watermark(canvas, doc):
            if settings['show_watermark']:
                canvas.saveState()
                canvas.setFont('Helvetica-Bold', 60)
                canvas.setFillColor(colors.lightgrey, alpha=0.3)
                canvas.translate(doc.width / 2.0, doc.height / 2.0)
                canvas.rotate(45)
                canvas.drawCentredString(0, 0, school_profile.get('watermark_text', 'CONFIDENTIAL'))
                canvas.restoreState()

        # Main content frame
        frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height - 0.95*inch, # Adjusted for report-card header
                      leftPadding=0, bottomPadding=0, rightPadding=0, topPadding=0,
                      showBoundary=0)
        main_template = PageTemplate(id='main_page', frames=[frame], onPage=_header_footer)
        doc.addPageTemplates([main_template])

        elements = [] # Initialize elements list for content
        # Build actual content (watermark will be applied during build)
        _set_progress(progress, 20, "Building broadsheet table")

        # Main Broadsheet Table
        subjects = data['subjects']
        headers = ["Pos", "Adm No", "Name", "Sex"] + subjects + ["Tot", "Avg", "Pts", "Div"]
        
        table_data = [headers]
        for r in data['rows']:
            row_vals = [str(r['Position']), r['Admission No'], r['Student Name'], r['Gender'] if r['Gender'] else '-']
            for s in subjects:
                row_vals.append(str(r['marks'].get(s, '-'))) # Use .get for safety
            row_vals += [str(r['Total']), str(r['Average']), str(r['Points']), str(r['Division']) if r['Division'] else '-']
            table_data.append(row_vals)
        
        # Dynamic font size based on subject count
        # Adjust column widths dynamically
        col_widths = [0.4*inch, 0.8*inch, 1.5*inch, 0.4*inch] + [0.6*inch] * len(subjects) + [0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch]
        total_width = sum(col_widths)
        if total_width > landscape(A4)[0] - 1*inch: # If too wide, scale down
            scale_factor = (landscape(A4)[0] - 1*inch) / total_width
            col_widths = [w * scale_factor for w in col_widths]

        font_size = 8
        if len(subjects) > 10: font_size = 7
        if len(subjects) > 15: font_size = 6
        
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), accent_dark),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size + 1), # Header font slightly larger
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"<b>BROADSHEET FOR {meta['class']} ({meta['level']})</b>", styles['SectionHeader']))
        elements.append(Spacer(1, 0.1*inch))

        elements.append(t)
        elements.append(PageBreak()) # Start analytics on a new page
        _set_progress(progress, 45, "Building analysis sections")

        # 1. Class Performance Analysis
        class_perf = data['class_performance']
        elements.append(Paragraph("CLASS PERFORMANCE ANALYSIS", styles['SectionHeader']))
        class_perf_data = [
            ["Total Students:", class_perf['total_students'], "Class Average:", f"{class_perf['class_average']}%"],
            ["Highest Average:", f"{class_perf['highest_average']}%", "Lowest Average:", f"{class_perf['lowest_average']}%"],
            ["Pass Rate:", f"{class_perf['pass_rate']}%", "Fail Rate:", f"{class_perf['fail_rate']}%"]
        ]
        class_perf_table = Table(class_perf_data, colWidths=[2*inch, 1.5*inch, 2*inch, 1.5*inch])
        class_perf_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('RIGHTPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(class_perf_table)
        elements.append(Spacer(1, 0.2*inch))

        # 2. Gender Summary
        if settings['show_gender_summary']:
            gender_sum = data['gender_summary']
            elements.append(Paragraph("GENDER SUMMARY", styles['SectionHeader']))
            gender_data = [["Gender", "Count"], ["Male", gender_sum['Male']], ["Female", gender_sum['Female']], ["Total", gender_sum['Total']]]
            gender_table = Table(gender_data, colWidths=[2*inch, 1.5*inch])
            gender_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            elements.append(gender_table)
            elements.append(Spacer(1, 0.2*inch))

        # 3. Division Summary
        div_sum = data['division_summary']
        elements.append(Paragraph("DIVISION SUMMARY", styles['SectionHeader']))
        div_data = [["Division", "Students"]]
        for div, count in div_sum.items():
            div_data.append([div, count])
        div_table = Table(div_data, colWidths=[2*inch, 1.5*inch])
        div_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), light),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(div_table)
        elements.append(Spacer(1, 0.2*inch))

        # 4. Top 10 Students
        top_students = data['top_students']
        elements.append(Paragraph("TOP 10 STUDENTS", styles['SectionHeader']))
        top_data = [["Position", "Admission No", "Student Name", "Average", "Division"]]
        for s in top_students:
            top_data.append([_pick(s, 'position', 'Position'), _pick(s, 'admission', 'Admission No'), _pick(s, 'name', 'Student Name'), _pick(s, 'average', 'Average'), _pick(s, 'division', 'Division')])
        top_table = Table(top_data, colWidths=[0.8*inch, 1.2*inch, 2.5*inch, 1*inch, 1*inch])
        top_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), light),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(top_table)
        elements.append(Spacer(1, 0.2*inch))

        # 5. Bottom 10 Students
        bottom_students = data['bottom_students']
        elements.append(Paragraph("BOTTOM 10 STUDENTS", styles['SectionHeader']))
        bottom_data = [["Position", "Admission No", "Student Name", "Average", "Division"]]
        for s in bottom_students:
            bottom_data.append([_pick(s, 'position', 'Position'), _pick(s, 'admission', 'Admission No'), _pick(s, 'name', 'Student Name'), _pick(s, 'average', 'Average'), _pick(s, 'division', 'Division')])
        bottom_table = Table(bottom_data, colWidths=[0.8*inch, 1.2*inch, 2.5*inch, 1*inch, 1*inch])
        bottom_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), light),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(bottom_table)
        elements.append(Spacer(1, 0.2*inch))

        # 6. Subject Performance Analysis
        sub_perf = data['subject_performance']
        elements.append(Paragraph("SUBJECT PERFORMANCE ANALYSIS", styles['SectionHeader']))
        sub_perf_data = [["Subject", "Average", "Passes", "Fails"]]
        for sub_name, stats in sub_perf.items():
            sub_perf_data.append([sub_name, stats['average'], stats['passes'], stats['fails']])
        sub_perf_table = Table(sub_perf_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch])
        sub_perf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), light),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(sub_perf_table)
        elements.append(Paragraph(f"Best Subject: {data['best_subject']} (Avg: {data['max_avg']})", styles['Normal']))
        elements.append(Paragraph(f"Worst Subject: {data['worst_subject']} (Avg: {data['min_avg']})", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

        # 7. Subject Ranking
        if settings['show_subject_ranking']:
            sub_ranking = data['subject_ranking']
            elements.append(Paragraph("SUBJECT RANKING (by Average Score)", styles['SectionHeader']))
            sub_ranking_data = [["Rank", "Subject"]]
            for rank, (sub_name, stats) in enumerate(sub_ranking, 1):
                sub_ranking_data.append([rank, f"{sub_name} (Avg: {stats['average']})"])
            sub_ranking_table = Table(sub_ranking_data, colWidths=[1*inch, 3*inch])
            sub_ranking_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            elements.append(sub_ranking_table)
            elements.append(Spacer(1, 0.2*inch))

        # Signatures
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph("Academic Master Signature: _________________________<br/><br/>", styles['Normal']))
        elements.append(Paragraph("Head Teacher Signature: _________________________<br/><br/>", styles['Normal']))
        elements.append(Paragraph("School Stamp:", styles['Normal']))

        # Build PDF and apply watermark callbacks (header/footer applied via PageTemplate)
        _set_progress(progress, 80, "Rendering PDF pages")
        doc.build(elements, onFirstPage=draw_watermark, onLaterPages=draw_watermark)
        _set_progress(progress, 100, "Export complete")
        QMessageBox.information(parent, "Success", f"Broadsheet exported to {path}")
    except Exception as e:
        print(f"[ERROR] Broadsheet export failed: {e}")
        QMessageBox.critical(parent, "Export Error", "An unexpected error occurred during export.")
    finally:
        if progress is not None:
            progress.close()
