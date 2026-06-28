import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import random
import os
import shutil

def create_styled_excel(filename, title, headers, data, sample=None):
    """Generates an Excel file matching the professional SRMS V5 template style."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # 1. School Header (Rows 1-3)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="STATE UNIVERSITY DEMONSTRATION SECONDARY SCHOOL").font = Font(size=16, bold=True)
    ws.cell(row=1, column=1).alignment = center_align

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value="Tanzania | srms@example.com").font = Font(size=10)
    ws.cell(row=2, column=1).alignment = center_align

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws.cell(row=3, column=1, value=title.upper()).font = Font(size=14, bold=True, color="2563EB")
    ws.cell(row=3, column=1).alignment = center_align

    # 2. Instructions (Rows 5-9)
    ws.cell(row=5, column=1, value="INSTRUCTIONS:").font = bold_font
    instructions = [
        "1. Do not modify the column headers in Row 10.",
        "2. Start data entry from Row 12 (below the sample row).",
        "3. Ensure the 'Admission No' exists in the system where required."
    ]
    for i, text in enumerate(instructions):
        ws.cell(row=6 + i, column=1, value=text).font = Font(italic=True, size=9)

    # 3. Headers (Row 10)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num, value=header)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_align

    # 4. Sample row (Row 11)
    if sample:
        for col_num, val in enumerate(sample, 1):
            cell = ws.cell(row=11, column=col_num, value=val)
            cell.font = Font(italic=True, color="808080")

    # 5. Data (Row 12+)
    for r_idx, row_data in enumerate(data, 12):
        for c_idx, value in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass  # cell.value may be None; skip width calculation
        ws.column_dimensions[column_letter].width = max_length + 4

    wb.save(filename)

def generate_pack():
    pack_dir = "SRMS_V5_TEST_PACK"
    if os.path.exists(pack_dir):
        shutil.rmtree(pack_dir)
    os.makedirs(pack_dir)

    # Realistic Tanzanian Names
    first_names = ["Juma", "Bakari", "Ali", "Hassan", "Said", "Hamisi", "Ramadhani", "Neema", "Rehema", "Asha", "Fatuma", "Maryam", "Amina", "Peter", "John", "Emmanuel", "Godfrey", "Salum", "Khalid", "Grace", "Anna", "Hellen", "Sophia"]
    last_names = ["Mbowe", "Kikwete", "Magufuli", "Mkapa", "Nyerere", "Karume", "Mwinyi", "Majaliwa", "Kassim", "Msigwa", "Lema", "Lissu", "Zitto", "Kibacho", "Chalamila", "Makamba", "Mchengerwa"]
    
    classes = [
        ("Form I", 100, "O_LEVEL"), ("Form II", 100, "O_LEVEL"),
        ("Form III", 100, "O_LEVEL"), ("Form IV", 100, "O_LEVEL"),
        ("Form V", 120, "A_LEVEL")
    ]
    
    o_subjects = ["Mathematics", "English", "Kiswahili", "Biology", "Chemistry", "Physics", "History", "Geography", "Civics"]
    a_subjects = ["History", "Geography", "Economics", "General Studies"]

    # 1. Students
    students_data = []
    o_level_adm = []
    a_level_adm = []
    adm_counter = 1
    for class_name, count, level in classes:
        for _ in range(count):
            adm = f"2024/{adm_counter:04d}"
            name = f"{random.choice(first_names)} {random.choice(last_names)}"
            gender = random.choice(["Male", "Female"])
            stream = random.choice(["A", "B", "C"])
            students_data.append([adm, name, gender, class_name, stream, level])
            if level == "O_LEVEL": o_level_adm.append(adm)
            else: a_level_adm.append(adm)
            adm_counter += 1
    
    create_styled_excel(f"{pack_dir}/students_import.xlsx", "STUDENT REGISTRATION FORM", 
                       ["Admission No", "Full Name", "Gender", "Class", "Stream", "Level"], 
                       students_data, ["2024/0001", "John Doe", "Male", "Form I", "A", "O_LEVEL"])

    # 2. Subjects
    subjects_data = []
    for s in o_subjects: subjects_data.append([s, "COUNTED", "O_LEVEL"])
    for s in a_subjects:
        stype = "SUBSIDIARY" if s == "General Studies" else "PRINCIPAL"
        subjects_data.append([s, stype, "A_LEVEL"])
    
    create_styled_excel(f"{pack_dir}/subjects_import.xlsx", "SUBJECT REGISTRATION FORM", 
                       ["Subject Name", "Subject Type", "Level"], 
                       subjects_data, ["Mathematics", "COUNTED", "O_LEVEL"])

    # 3. Enrollments
    enroll_data = []
    for adm in o_level_adm:
        for s in o_subjects: enroll_data.append([adm, s])
    for adm in a_level_adm:
        for s in a_subjects: enroll_data.append([adm, s])
    
    create_styled_excel(f"{pack_dir}/enrollments_import.xlsx", "STUDENT SUBJECT ENROLLMENT FORM", 
                       ["Admission No", "Subject Name"], 
                       enroll_data, ["2024/0001", "Mathematics"])

    # 4. Requirements
    req_items = ["Ream Paper", "Laboratory Contribution", "Examination Fee", "Identity Card", "Sports Contribution", "Practical Fee", "School Development Fund"]
    req_data = [[item, random.randint(1, 5), "Per Term"] for item in req_items]
    create_styled_excel(f"{pack_dir}/requirements_import.xlsx", "SCHOOL REQUIREMENTS TEMPLATE", 
                       ["Item Name", "Quantity", "Notes"], 
                       req_data, ["Reams of Paper", "2", "Urgent"])

    # 5. Results (O-Level and A-Level)
    for sub in o_subjects:
        res = [[adm, random.randint(30, 95)] for adm in o_level_adm]
        create_styled_excel(f"{pack_dir}/{sub.lower()}_results.xlsx", f"{sub.upper()} RESULTS", ["Admission No", "Marks"], res)

    for sub in a_subjects:
        res = [[adm, random.randint(20, 90)] for adm in a_level_adm]
        fn = sub.lower().replace(" ", "_") + "_al_results.xlsx"
        create_styled_excel(f"{pack_dir}/{fn}", f"{sub.upper()} RESULTS", ["Admission No", "Marks"], res)

    shutil.make_archive(pack_dir, 'zip', pack_dir)
    print(f"Generated {pack_dir}.zip successfully.")

if __name__ == "__main__":
    generate_pack()