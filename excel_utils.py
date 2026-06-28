import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from PySide6.QtWidgets import QFileDialog, QMessageBox
from db_utils import fetch_one
import os

def download_template(parent, filename, title, headers, instructions=None, samples=None):
    path, _ = QFileDialog.getSaveFileName(parent, "Download Template", filename, "Excel Files (*.xlsx)")
    if not path: return
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"

        # 1. School Header
        profile = fetch_one("SELECT school_name, school_address, school_phone, school_email FROM school_profile LIMIT 1")

        school_name = profile[0].upper() if profile and profile[0] else "SCHOOL MANAGEMENT SYSTEM"
        school_contact = f"{profile[1] if profile and profile[1] else '-'} | {profile[2] if profile and profile[2] else '-'} | {profile[3] if profile and profile[3] else '-'}"

        # Styling
        blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Row 1-3: School Info & Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=school_name).font = Font(size=16, bold=True)
        ws.cell(row=1, column=1).alignment = center_align

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1, value=school_contact).font = Font(size=10)
        ws.cell(row=2, column=1).alignment = center_align

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        ws.cell(row=3, column=1, value=title.upper()).font = Font(size=14, bold=True, color="2563EB")
        ws.cell(row=3, column=1).alignment = center_align

        # Row 5: Instructions
        ws.cell(row=5, column=1, value="INSTRUCTIONS:").font = bold_font
        if not instructions:
            instructions = [
                "1. Do not modify the column headers in Row 10.",
                "2. Start data entry from Row 12 (below the sample row).",
                "3. Ensure the 'Admission No' exists in the system where required."
            ]
        for i, text in enumerate(instructions):
            ws.cell(row=6 + i, column=1, value=text).font = Font(italic=True, size=9)

        # Row 10: Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col_num, value=header)
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_align

        # Row 11: Sample Row
        if samples:
            for col_num, val in enumerate(samples, 1):
                cell = ws.cell(row=11, column=col_num, value=val)
                cell.font = Font(italic=True, color="808080")

        # Freeze Panes (Header stays visible)
        ws.freeze_panes = "A11"

        # Auto Column Width
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            ws.column_dimensions[column_letter].width = max_length + 4

        wb.save(path)
        QMessageBox.information(parent, "Success", f"Template saved to {path}")
    except Exception as e:
        QMessageBox.critical(parent, "Error", "Failed to save template. Please check the file path and try again.")

def export_to_excel(parent, filename, headers, data):
    path, _ = QFileDialog.getSaveFileName(parent, "Export Data", filename, "Excel Files (*.xlsx)")
    if not path: return
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data:
            ws.append(row)
        wb.save(path)
        QMessageBox.information(parent, "Success", f"Data exported to {path}")
    except Exception as e:
        QMessageBox.critical(parent, "Error", "Failed to export data. Please check the file path and try again.")

def get_import_file(parent):
    path, _ = QFileDialog.getOpenFileName(parent, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
    return path
